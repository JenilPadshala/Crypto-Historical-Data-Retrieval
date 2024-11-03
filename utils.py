from datetime import datetime
from typing import Optional, Tuple
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score


###########     helper functions for main.py     ############


def add_to_excel(data: pd.DataFrame, sheet_name: str) -> None:
    """
    Add the DataFrame to an Excel file with the specified sheet name.

    Parameters:
    data (pd.DataFrame): DataFrame to be added to the Excel file.
    sheet_name (str): Name of the sheet in the Excel file.
    """
    sheet_name = sheet_name.replace("/", "")

    filepath = "output.xlsx"
    try:
        book = load_workbook(filepath)
        if sheet_name in book.sheetnames:
            # Append a suffix to the sheet name to make it unique
            suffix = 1
            new_sheet_name = f"{sheet_name}_{suffix}"
            while new_sheet_name in book.sheetnames:
                suffix += 1
                new_sheet_name = f"{sheet_name}_{suffix}"
            sheet_name = new_sheet_name
    except FileNotFoundError:
        # File does not exist, no need to check for sheet names
        pass

    with pd.ExcelWriter(
        filepath, engine="openpyxl", mode="a" if "book" in locals() else "w"
    ) as writer:
        data.to_excel(writer, sheet_name=sheet_name, index=False)


###########     helper functions for data_retrieval.py     ############
def date_str_to_unix_timestamp(date: str) -> int:
    """
    Convert a date string to a Unix timestamp

    Args:
    date (str): The date in the format 'YYYY-MM-DD'

    Returns:
    int: The Unix timestamp for the given date
    """
    return int(datetime.strptime(date, "%Y-%m-%d").timestamp())


def filter_data_by_date(data: pd.DataFrame, start_date: str) -> pd.DataFrame:
    """
    Filter the data by the start date.

    Parameters:
    data (pd.DataFrame): DataFrame containing cryptocurrency data.
    start_date (str): Start date to filter the data.

    Returns:
    pd.DataFrame: Filtered DataFrame with data starting from the specified date.
    """
    return data[data["Date"] >= start_date]


def parse_response_data(data: list) -> pd.DataFrame:
    """
    Parse JSON response data into a DataFrame with columns ['Date', 'Open', 'High', 'Low', 'Close'].

    Parameters:
    data (list): List of data items from the API response.

    Returns:
    pd.DataFrame: Parsed DataFrame with cryptocurrency data.
    """
    records = [
        {
            "Date": datetime.utcfromtimestamp(item["TIMESTAMP"]),
            "Open": item.get("OPEN"),
            "High": item.get("HIGH"),
            "Low": item.get("LOW"),
            "Close": item.get("CLOSE"),
        }
        for item in data
    ]
    return pd.DataFrame(records)


###########     helper functions for metrics_calculations.py     ############
def calculate_historical_high(data: pd.DataFrame, days: int) -> pd.Series:
    """Calculate the historical high over the past `days` days."""

    return data["High"].rolling(window=days, min_periods=1).max()


def calculate_historical_low(data: pd.DataFrame, days: int) -> pd.Series:
    """Calculate the historical low over the past `days` days."""

    return data["Low"].rolling(window=days, min_periods=1).min()


def days_since_high(data: pd.DataFrame, index: int, days: int) -> Optional[int]:
    """Calculate the number of days since the historical high over the past `days` days."""

    subset = data.loc[max(0, index - days + 1) : index]
    high_value = subset["High"].max()
    high_date = subset.loc[subset["High"] == high_value, "Date"].max()
    return (
        (data.loc[index, "Date"] - high_date).days if high_date is not pd.NaT else None
    )


def days_since_low(data: pd.DataFrame, index: int, days: int) -> Optional[int]:
    """Calculate the number of days since the historical low over the past `days` days."""

    subset = data.loc[max(0, index - days + 1) : index]
    low_value = subset["Low"].min()
    low_date = subset.loc[subset["Low"] == low_value, "Date"].max()
    return (data.loc[index, "Date"] - low_date).days if low_date is not pd.NaT else None


def calculate_pct_diff(current: pd.Series, reference: pd.Series) -> pd.Series:
    """Calculate the percentage difference between two series."""

    return round(((current - reference) / reference) * 100, 2)


def calculate_future_high(data: pd.DataFrame, days: int) -> pd.Series:
    """Calculate the future high over the next `days` days."""

    future_high = [
        (
            data["High"].iloc[i + 1 : min(len(data), i + days)].max()
            if i < len(data) - 1
            else data["High"].iloc[i]
        )
        for i in range(len(data))
    ]
    return pd.Series(future_high, index=data.index)


def calculate_future_low(data: pd.DataFrame, days: int) -> pd.Series:
    """Calculate the future low over the next `days` days."""

    future_low = [
        (
            data["Low"].iloc[i + 1 : min(len(data), i + days)].min()
            if i < len(data) - 1
            else data["Low"].iloc[i]
        )
        for i in range(len(data))
    ]
    return pd.Series(future_low, index=data.index)


###########     helper functions for ml_model.py     ############
def evaluation_metrics(
    y_true: np.ndarray, y_pred: np.ndarray
) -> Tuple[float, float, float]:
    """
    Calculates evaluation metrics for the model predictions.

    Parameters:
    y_true (np.ndarray): The true target values.
    y_pred (np.ndarray): The predicted target values.

    Returns:
    Tuple[float, float, float]: A tuple containing the Mean Squared Error (MSE), Mean Absolute Error (MAE), and R-squared (R2) score.
    """
    mse = mean_squared_error(y_true, y_pred)
    mae = mean_absolute_error(y_true, y_pred)
    r2 = r2_score(y_true, y_pred)
    return mse, mae, r2
